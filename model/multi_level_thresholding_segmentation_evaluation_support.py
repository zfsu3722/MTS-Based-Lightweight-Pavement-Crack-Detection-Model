import numpy as np
import crack500_support as cr5_spt
import crack_self_proc as cr_sf_proc
import matplotlib.pyplot as plt
import time as tm
import img_quality_metrics as img_qm
import cv2
import self_imq_implement_support as imq_spt
import otsu_obj_fun as otsu_obj
import equilibrium_optimizer as EQOPT
import hbo_optimizer as hbo_opt
from otsu_obj_fun import PSOLsp
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

IS_SEGMENT_MOD_PARAM = "IS_SEGMENT_MOD"
IS_IMG_INT_PARAM = "IS_IMG_INT"
OPT_OBJ_FUN_PARAM = "OPT_OBJ_FUN"
EVAL_TRIAL_NUM_PARAM = "EVAL_TRIAL_NUM_PARAM"
EVAL_SEG_METHOD_PARAM = "EVAL_SEG_METHOD_PARAM"
EVAL_SEG_METHOD_NAME_PARAM = "EVAL_SEG_METHOD_NAME_PARAM"
EVAL_RESULT_DUMP_FILE_PATH_PARAM = "EVAL_RESULT_DUMP_FILE_PATH_PARAM"
EVAL_METRIC_LIST_PARAM = "EVAL_METRIC_LIST_PARAM"
EVAL_METRIC_KEY_LIST_PARAM = "EVAL_METRIC_KEY_LIST_PARAM"
EVAL_SEG_METHOD_NAME_KEY = "EVAL_SEG_METHOD_NAME"
EVAL_STATISTICS_MEAN_KEY = "EVAL_STATISTICS_MEAN"
EVAL_STATISTICS_STD_KEY = "EVAL_STATISTICS_STD"
EVAL_SEG_METHOD_INTEGER = "EQUAL_RANGE"
EVAL_SEG_METHOD_HBO = "HBO"
EVAL_SEG_METHOD_PSO = "PSO"
EVAL_SEG_METHOD_EQB = "EQB"
EVAL_HIST_CLASS_PARAM = "EVAL_HIST_CLASS_PARAM"
EVAL_HIST_CLASS_KEY = "EVAL_CLASS"
EVAL_HIST_CLASS_1 = "SINGLE"
EVAL_HIST_CLASS_2 = "DOUBLE"
EVAL_HIST_CLASS_3 = "TRIPLE"
EVAL_HIST_CLASS_4 = "LARGE_AREA"
EVAL_THRESHOLD_LIST_KEY = "THRESHOLD_LIST"
EVAL_IMG_PLANE_NAME_LIST_KEY = "IMG_PLANE_NAME_LIST"
FSIM_KEY = "FSIM"
SSIM_KEY = "SSIM"
PSNR_KEY = "PSNR"
RMSE_KEY = "RMSE"
SEG_TIME_KEY = "SEG_TIME"
IMG_PLANE_MAX = 255
EVAL_METRIC_KEY_IDX = 0
EVAL_METRIC_METHOD_IDX = 1
EVAL_CURVE_NAME_IDX = 0
EVAL_CURVE_THRESHOLD_IDX = 1
EVAL_CURVE_MEAN_IDX = 2
EVAL_CURVE_STD_IDX = 3
EVAL_METRIC_NAME_KEY = 0
EVAL_METRIC_VALUE_KEY = 1
EVAL_METRIC_TABLE_QUERY_RESULT_CLASS_TYPE_IDX = 0
EVAL_METRIC_TABLE_QUERY_METRIC_TYPE_IDX = 1
EVAL_METRIC_TABLE_QUERY_SEG_METHOD_IDX = 2
EVAL_METRIC_TABLE_QUERY_THRESHOLD_NUM_IDX = 3
EVAL_METRIC_TABLE_QUERY_IMAGE_PLANE_IDX = 4
EVAL_METRIC_TABLE_QUERY_METRIC_VALUE_IDX = 5
EVAL_METRIC_TABLE_QUERY_METRIC_MEAN_VALUE_IDX = 0
EVAL_METRIC_TABLE_QUERY_METRIC_STD_VALUE_IDX = 1
EVAL_METRIC_TABLE_QUERY_SHEET_NAME_CONN = '_'
XLS_SHEET_NAME_INITIAL = "Sheet"
XLS_SHEET_INITIAL_NO_FOUND_MSG = "No Sheet Found"
EVAL_METRIC_TABLE_QUERY_METRIC_SEG_METHOD_COLUMN_START = 3
EVAL_METRIC_TABLE_QUERY_METRIC_SEG_METHOD_ROW = 1
EVAL_METRIC_TABLE_QUERY_METRIC_THRESHOLD_COLUMN = 1
EVAL_METRIC_TABLE_QUERY_METRIC_IMAGE_PLANE_NAME_COLUMN = 2
EVAL_METRIC_XLS_THRESHOLD_NUM_TITLE = "Threshold Number"
EVAL_METRIC_XLS_IMAGE_NAME_TITLE = "Image Name"
EVAL_METRIC_XLS_TITLE_ROW = 1
EVAL_METRIC_XLS_TITLE_COL_1 = 1
EVAL_METRIC_XLS_TITLE_COL_2 = 2
EVAL_METRIC_XLS_DATA_ROW_START = 2
FILE_NAME_EXT_SPLITER = "."
CELL_BOARDER = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))


def img_threshold_reconstruction_plane(img_plane, threshold_list):
    segmentation_img_plane_list = cr5_spt.img_segmentation_threshold_list_light(img_plane, threshold_list)
    img_plane_rec = cr5_spt.img_gray_planes_reconstruction_avg_mask(img_plane, segmentation_img_plane_list)
    return img_plane_rec


def img_multi_level_segmentation_reconstruction_integer_plane(img_plane, threshold_num, additional_params):
    is_segment_mod = additional_params[IS_SEGMENT_MOD_PARAM]
    is_img_int = additional_params[IS_IMG_INT_PARAM]
    if is_img_int:
        img_plane = img_plane / IMG_PLANE_MAX
    seg_num = threshold_num + 1
    seg_start_time = tm.time()
    multi_threshold_integer_float, multi_threshold_integer_int, threshold_list_with_zero, threshold_list_with_zero_int, plane_section_num, effective_threshold_last = cr5_spt.img_integer_segmentation_equal_range_thresholds_light(img_plane, seg_num, is_segment_mod=is_segment_mod)
    seg_end_time = tm.time()
    seg_time = seg_end_time - seg_start_time
    img_plane_rec = img_threshold_reconstruction_plane(img_plane, multi_threshold_integer_int / IMG_PLANE_MAX)
    img_plane_rec = np.round(img_plane_rec*IMG_PLANE_MAX, 0).astype(np.uint8)
    return img_plane_rec, seg_time


def eqb_opt_object_fun(img_plane, threshold_num):
    opt_problem = otsu_obj.AGDivergence(img_plane, threshold_num)
    seg_obj = EQOPT.EO(problem=opt_problem, popSize=100, maxIter=500, parameters=EQOPT.parameters)
    return seg_obj


def hbo_opt_object_fun(img_plane, threshold_num):
    l_b = np.ones(threshold_num)
    u_b = np.ones(threshold_num) * IMG_PLANE_MAX
    otsu_fun = otsu_obj.OtsuObjFun(img_plane, threshold_num)
    func_obj = otsu_fun.seg_obj
    seg_obj = hbo_opt.HBO(searchAgents=40, Max_iter=100, lb=l_b, ub=u_b, dim=threshold_num, fobj=func_obj, cycles=10, degree=5)
    return seg_obj


def pso_opt_object_fun(img_plane, threshold_num):
    l_b = np.ones(threshold_num)
    u_b = np.ones(threshold_num) * IMG_PLANE_MAX
    otsu_fun = otsu_obj.OtsuObjFun(img_plane, threshold_num)
    func_obj = otsu_fun.seg_obj
    seg_obj = PSOLsp(func=func_obj, n_dim=threshold_num, pop=40, max_iter=3000, lb=l_b, ub=u_b, w=0.8, c1=0.5, c2=0.5, lsp_on=False, lsp_prob=0.5)
    return seg_obj


OPT_OBJ_FUN_HBO = hbo_opt_object_fun
OPT_OBJ_FUN_EQB = eqb_opt_object_fun
OPT_OBJ_FUN_PSO = pso_opt_object_fun


def img_multi_level_segmentation_reconstruction_opt(img_plane, threshold_num, additional_params):
    is_img_int = additional_params[IS_IMG_INT_PARAM]
    opt_object_fun = additional_params[OPT_OBJ_FUN_PARAM]
    if is_img_int is not True:
        img_plane = np.round(img_plane * IMG_PLANE_MAX, 0).astype(np.uint8)
    seg_obj = opt_object_fun(img_plane, threshold_num)
    seg_start_time = tm.time()
    seg_obj.run()
    seg_end_time = tm.time()
    seg_time = seg_end_time - seg_start_time
    g_best_x = seg_obj.get_best_x()
    threshold_list = cr5_spt.opt_result_to_threshold_list(g_best_x, is_img_int=True)
    img_plane_rec = img_threshold_reconstruction_plane(img_plane, threshold_list)
    img_plane_rec = np.round(img_plane_rec, 0).astype(np.uint8)
    return img_plane_rec, seg_time


EVAL_SEG_INTEGER = img_multi_level_segmentation_reconstruction_integer_plane
EVAL_SEG_OPT = img_multi_level_segmentation_reconstruction_opt


def eval_img_plane_int_type_sim(img_plane_ref, img_plane_dist):
    fsim = imq_spt.fsim(img_plane_ref, img_plane_dist)
    ssim = img_qm.ssim(img_plane_ref, img_plane_dist, max_p=IMG_PLANE_MAX)
    psnr = img_qm.psnr(img_plane_ref, img_plane_dist, max_p=IMG_PLANE_MAX)
    rmse = img_qm.rmse(img_plane_ref, img_plane_dist, max_p=IMG_PLANE_MAX)
    return fsim, ssim, psnr, rmse


def eval_img_plane_sim_multi_times(img_plane, threshold_num, additional_params):
    trial_num = additional_params[EVAL_TRIAL_NUM_PARAM]
    seg_method = additional_params[EVAL_SEG_METHOD_PARAM]
    is_img_int = additional_params[IS_IMG_INT_PARAM]
    if is_img_int is not True:
        img_plane_int = np.round(img_plane * IMG_PLANE_MAX, 0).astype(np.uint8)
    else:
        img_plane_int = img_plane
    fsim_result_list = []
    ssim_result_list = []
    psnr_result_list = []
    rmse_result_list = []
    seg_time_list = []
    eval_result = dict()
    i = 0
    while i < trial_num:
        img_plane_rec, seg_time = seg_method(img_plane, threshold_num, additional_params)
        fsim, ssim, psnr, rmse = eval_img_plane_int_type_sim(img_plane_int, img_plane_rec)
        fsim_result_list.append(cr5_spt.nan_proc_replace(fsim))
        ssim_result_list.append(cr5_spt.nan_proc_replace(ssim))
        psnr_result_list.append(cr5_spt.nan_proc_replace(psnr))
        rmse_result_list.append(cr5_spt.nan_proc_replace(rmse))
        seg_time_list.append(seg_time)
        print("trial ", i, "completed")
        i += 1
    eval_result[FSIM_KEY] = np.array(fsim_result_list)
    eval_result[SSIM_KEY] = np.array(ssim_result_list)
    eval_result[PSNR_KEY] = np.array(psnr_result_list)
    eval_result[RMSE_KEY] = np.array(rmse_result_list)
    eval_result[SEG_TIME_KEY] = np.array(seg_time_list)
    return eval_result


def eval_img_plane_list_single_threshold_num(img_plane_list, img_plane_name_list, threshold_num, additional_params):
    eval_result = dict()
    plane_num = len(img_plane_list)
    i = 0
    while i < plane_num:
        img_plane = img_plane_list[i]
        img_plane_key = img_plane_name_list[i]
        print("threshold_num ", threshold_num, "plane_num", i, "started")
        #sim_eval_result = eval_img_plane_sim_multi_times(img_plane, threshold_num, additional_params)
        sim_eval_result = eval_img_plane_sim_multi_times_dynamic(img_plane, threshold_num, additional_params)
        eval_result[img_plane_key] = sim_eval_result
        #print("threshold_num ", threshold_num, "plane_num", i, "completed")
        i += 1
    return eval_result


def eval_img_plane_list_multi_threshold_num(img_plane_list, img_plane_name_list, threshold_num_list, additional_params):
    eval_hist_class = additional_params[EVAL_HIST_CLASS_PARAM]
    eval_method_name = additional_params[EVAL_SEG_METHOD_NAME_PARAM]
    eval_result = dict()
    eval_result[EVAL_HIST_CLASS_KEY] = eval_hist_class
    eval_result[EVAL_SEG_METHOD_NAME_KEY] = eval_method_name
    eval_result[EVAL_THRESHOLD_LIST_KEY] = np.array(threshold_num_list)
    eval_result[EVAL_IMG_PLANE_NAME_LIST_KEY] = img_plane_name_list
    threshold_total = len(threshold_num_list)
    i = 0
    while i < threshold_total:
        threshold_num = threshold_num_list[i]
        threshold_eval_result = eval_img_plane_list_single_threshold_num(img_plane_list, img_plane_name_list, threshold_num, additional_params)
        threshold_num_key = str(threshold_num)
        eval_result[threshold_num_key] = threshold_eval_result
        i += 1
    eval_result_dump_file_path = additional_params[EVAL_RESULT_DUMP_FILE_PATH_PARAM]
    if eval_result_dump_file_path is not None:
        eval_result_dump(eval_result_dump_file_path, eval_result)
    return eval_result


def eval_result_dump(dump_file_path, eval_result):
    cr5_spt.package_img_file(dump_file_path, eval_result)
    return


def load_eval_img_plane(img_class_path, img_plane_idx, is_img_int=False):
    file_path_list, img_plane_name_list = cr5_spt.retrieve_file_paths(img_class_path)
    img_plane_list = cr5_spt.retrieve_mos_raw_images(file_path_list, is_rgb=False, img_map_idx=img_plane_idx, is_img_int=is_img_int)
    return img_plane_list, img_plane_name_list


def analyse_eval_metrics_mean_curve_statistics(eval_metric_file_path, additional_params, eval_statistics_result_file_path=None):
    eval_metric_result_list = load_eval_metrics(eval_metric_file_path)
    eval_metric_statistics_result = extract_eval_metric_statistics(eval_metric_result_list, additional_params)
    if eval_statistics_result_file_path is not None:
        eval_result_dump(eval_statistics_result_file_path, eval_metric_statistics_result)
    return eval_metric_statistics_result


def init_result_map_by_key_list(key_list, result_map=None):
    if result_map is None:
        result_map = dict()
    key_num = len(key_list)
    i = 0
    while i < key_num:
        key = key_list[i]
        result_map[str(key)] = dict()
        i += 1
    return result_map


def analyse_eval_metrics_statistics_image_wise_table(eval_metric_file_path, additional_params, eval_statistics_result_file_path=None):
    eval_metric_result_list = load_eval_metrics(eval_metric_file_path)
    eval_metric_statistics_result = extract_eval_metric_statistics_image_wise_table(eval_metric_result_list, additional_params)
    if eval_statistics_result_file_path is not None:
        eval_result_dump(eval_statistics_result_file_path, eval_metric_statistics_result)
    return eval_metric_statistics_result


def extract_eval_metric_statistics_image_wise_table(eval_metric_result_list, additional_params):
    eval_metric_key_list = additional_params[EVAL_METRIC_KEY_LIST_PARAM]
    eval_metric_statistics_result = init_eval_metric_statistics_result(eval_metric_result_list)
    eval_metric_result_num = len(eval_metric_result_list)
    i = 0
    while i < eval_metric_result_num:
        eval_metric_result = eval_metric_result_list[i]
        threshold_list = eval_metric_result[EVAL_THRESHOLD_LIST_KEY]
        image_plane_name_list = eval_metric_result[EVAL_IMG_PLANE_NAME_LIST_KEY]
        class_type = eval_metric_result[EVAL_HIST_CLASS_KEY]
        seg_method_name = eval_metric_result[EVAL_SEG_METHOD_NAME_KEY]
        if len(eval_metric_statistics_result[class_type]) == 1:
            eval_metric_statistics_result[class_type] = init_result_map_by_key_list(threshold_list, result_map=eval_metric_statistics_result[class_type])
            eval_metric_statistics_result[class_type][EVAL_IMG_PLANE_NAME_LIST_KEY] = image_plane_name_list
        eval_metric_statistics_result_class = eval_metric_statistics_result[class_type]
        eval_metric_statistics_result_class[EVAL_SEG_METHOD_NAME_KEY].append(seg_method_name)
        do_extract_eval_metric_statistics_image_wise_table(eval_metric_statistics_result_class, eval_metric_result, threshold_list, eval_metric_key_list, image_plane_name_list, seg_method_name)
        i += 1
    return eval_metric_statistics_result


def do_extract_eval_metric_statistics_image_wise_table(eval_metric_statistics_result_class, eval_metric_result, threshold_list, eval_metric_key_list, image_plane_name_list, seg_method_name):
    threshold_num = len(threshold_list)
    i = 0
    while i < threshold_num:
        threshold = threshold_list[i]
        eval_threshold_result = eval_metric_result[str(threshold)]
        if not eval_metric_statistics_result_class[str(threshold)]:
            eval_metric_statistics_result_class[str(threshold)] = init_result_map_by_key_list(eval_metric_key_list)
        eval_metric_statistics_result_threshold = eval_metric_statistics_result_class[str(threshold)]
        do_extract_eval_metric_statistics_image_wise_table_threshold(eval_metric_statistics_result_threshold, eval_threshold_result, eval_metric_key_list, image_plane_name_list, seg_method_name)
        i += 1
    return eval_metric_statistics_result_class


def do_extract_eval_metric_statistics_image_wise_table_threshold(eval_metric_statistics_result_threshold, eval_metric_result_threshold, eval_metric_key_list, image_plane_name_list, seg_method_name):
    image_plane_num = len(image_plane_name_list)
    i = 0
    while i < image_plane_num:
        img_plane_name = image_plane_name_list[i]
        eval_img_plane_result = eval_metric_result_threshold[img_plane_name]
        mean_result_list, std_result_list = do_extract_eval_metric_statistics_image_wise_table_image_plane(eval_img_plane_result, eval_metric_key_list)
        eval_metric_statistics_result_threshold = fill_eval_metric_statistics_image_wise_table_metric(eval_metric_statistics_result_threshold, img_plane_name, seg_method_name, mean_result_list, std_result_list)
        i += 1
    return eval_metric_statistics_result_threshold


def do_extract_eval_metric_statistics_image_wise_table_image_plane(eval_metric_result_image_plane, eval_metric_key_list):
    mean_result_list = []
    std_result_list = []
    metric_num = len(eval_metric_key_list)
    i = 0
    while i < metric_num:
        metric_key = eval_metric_key_list[i]
        v_metric_value = eval_metric_result_image_plane[metric_key]
        mean_result = (metric_key, np.mean(v_metric_value))
        std_result = (metric_key, np.std(v_metric_value))
        mean_result_list.append(mean_result)
        std_result_list.append(std_result)
        i += 1
    return mean_result_list, std_result_list


def fill_eval_metric_statistics_image_wise_table_metric(eval_metric_statistics_result_threshold, img_plane_name, seg_method_name, mean_result_list, std_result_list):
    metric_num = len(mean_result_list)
    i = 0
    while i < metric_num:
        metric_result_mean = mean_result_list[i]
        metric_result_std = std_result_list[i]
        metric_key = metric_result_mean[EVAL_METRIC_NAME_KEY]
        eval_metric_statistics_result_metric = eval_metric_statistics_result_threshold[metric_key]
        eval_metric_statistics_result_image_plane = eval_metric_statistics_result_metric.get(img_plane_name)
        if eval_metric_statistics_result_image_plane is None:
            eval_metric_statistics_result_metric[img_plane_name] = dict()
            eval_metric_statistics_result_image_plane = eval_metric_statistics_result_metric[img_plane_name]
        eval_metric_statistics_result_seg_method = eval_metric_statistics_result_image_plane.get(seg_method_name)
        if eval_metric_statistics_result_seg_method is None:
            eval_metric_statistics_result_image_plane[seg_method_name] = dict()
            eval_metric_statistics_result_seg_method = eval_metric_statistics_result_image_plane[seg_method_name]
        eval_metric_statistics_result_seg_method[EVAL_STATISTICS_MEAN_KEY] = metric_result_mean[EVAL_METRIC_VALUE_KEY]
        eval_metric_statistics_result_seg_method[EVAL_STATISTICS_STD_KEY] = metric_result_std[EVAL_METRIC_VALUE_KEY]
        i += 1
    return eval_metric_statistics_result_threshold


def query_eval_metric_statistics_image_wise_table_threshold_plane(eval_metric_table_file_path, class_type, threshold_list, metric_key, img_plane_name_list=None, seg_method_list=None, query_result_output_file_path=None):
    eval_metric_statistics_image_wise_table_result = cr5_spt.extract_img_objects(eval_metric_table_file_path)
    if seg_method_list is None:
        seg_method_list = eval_metric_statistics_image_wise_table_result[class_type][EVAL_SEG_METHOD_NAME_KEY]
    if img_plane_name_list is None:
        img_plane_name_list = eval_metric_statistics_image_wise_table_result[class_type][EVAL_IMG_PLANE_NAME_LIST_KEY]
    img_plane_num = len(img_plane_name_list)
    threshold_num = len(threshold_list)
    query_result_list = []
    i = 0
    while i < threshold_num:
        threshold = threshold_list[i]
        j = 0
        while j < img_plane_num:
            img_plane_name = img_plane_name_list[j]
            query_result = do_query_eval_metric_statistics_image_wise_table(eval_metric_statistics_image_wise_table_result, class_type, threshold, metric_key, img_plane_name, seg_method_list)
            query_result_list.append(query_result)
            j += 1
        i += 1
    query_result_sheet = (class_type, metric_key, seg_method_list, threshold_list, img_plane_name_list,  query_result_list)
    if query_result_output_file_path is not None:
        eval_result_dump(query_result_output_file_path, query_result_sheet)
    return query_result_sheet


def do_query_eval_metric_statistics_image_wise_table(eval_metric_statistics_image_wise_table_result, class_type, threshold, metric_key, img_plane_name, seg_method_list):
    eval_metric_statistics_image_wise_table_result_class = eval_metric_statistics_image_wise_table_result[class_type]
    eval_metric_statistics_image_wise_table_result_threshold = eval_metric_statistics_image_wise_table_result_class[str(threshold)]
    eval_metric_statistics_image_wise_table_result_metric = eval_metric_statistics_image_wise_table_result_threshold[metric_key]
    eval_metric_statistics_image_wise_table_result_img_plane = eval_metric_statistics_image_wise_table_result_metric[img_plane_name]
    seg_method_num = len(seg_method_list)
    metric_mean_list = []
    metric_std_list = []
    i = 0
    while i < seg_method_num:
        seg_method = seg_method_list[i]
        metric_result = eval_metric_statistics_image_wise_table_result_img_plane[seg_method]
        metric_mean_list.append(metric_result[EVAL_STATISTICS_MEAN_KEY])
        metric_std_list.append(metric_result[EVAL_STATISTICS_STD_KEY])
        i += 1
    query_result = (metric_mean_list, metric_std_list)
    return query_result


def load_eval_metrics(eval_metric_file_path):
    metric_file_path_list, metric_file_name_list = cr5_spt.retrieve_file_paths(eval_metric_file_path)
    eval_metric_result_list = extract_eval_metrics(metric_file_path_list)
    return eval_metric_result_list


def extract_eval_metrics(metric_file_path_list):
    metric_file_num = len(metric_file_path_list)
    eval_metric_result_list = []
    i = 0
    while i < metric_file_num:
        eval_metric_file_path = metric_file_path_list[i]
        eval_metric_result = cr5_spt.extract_img_objects(eval_metric_file_path)
        eval_metric_result_list.append(eval_metric_result)
        i += 1
    return eval_metric_result_list


def init_eval_metric_statistics_result(eval_metric_result_list):
    eval_metric_result_num = len(eval_metric_result_list)
    eval_metric_statistics_result = dict()
    i = 0
    while i < eval_metric_result_num:
        eval_metric_result = eval_metric_result_list[i]
        eval_class_type = eval_metric_result[EVAL_HIST_CLASS_KEY]
        if eval_class_type not in eval_metric_statistics_result:
            eval_metric_statistics_result[eval_class_type] = dict()
            eval_metric_statistics_result[eval_class_type][EVAL_SEG_METHOD_NAME_KEY] = []
        i += 1
    return eval_metric_statistics_result


def extract_eval_metric_statistics(eval_metric_result_list, additional_params):
    eval_metric_key_list = additional_params[EVAL_METRIC_KEY_LIST_PARAM]
    eval_metric_statistics_result = init_eval_metric_statistics_result(eval_metric_result_list)
    eval_metric_result_num = len(eval_metric_result_list)
    i = 0
    while i < eval_metric_result_num:
        eval_metric_result = eval_metric_result_list[i]
        seg_method_name = eval_metric_result[EVAL_SEG_METHOD_NAME_KEY]
        threshold_list = eval_metric_result[EVAL_THRESHOLD_LIST_KEY]
        eval_metric_result_statistics_mean, eval_metric_result_statistics_std = do_extract_eval_metric_statistics(eval_metric_result, eval_metric_key_list)
        eval_metric_statistics_result = fill_eval_metric_statistics_result(eval_metric_result, eval_metric_result_statistics_mean, eval_metric_result_statistics_std, eval_metric_statistics_result, threshold_list, seg_method_name)
        i += 1
    return eval_metric_statistics_result


def fill_eval_metric_statistics_result(eval_metric_result, eval_metric_result_statistics_mean, eval_metric_result_statistics_std, eval_metric_statistics_result, threshold_list, seg_method_name):
    eval_class_type = eval_metric_result[EVAL_HIST_CLASS_KEY]
    eval_seg_method_name = eval_metric_result[EVAL_SEG_METHOD_NAME_KEY]
    eval_metric_statistics_element = dict()
    eval_metric_statistics_element[EVAL_STATISTICS_MEAN_KEY] = eval_metric_result_statistics_mean
    eval_metric_statistics_element[EVAL_STATISTICS_STD_KEY] = eval_metric_result_statistics_std
    eval_metric_statistics_element[EVAL_THRESHOLD_LIST_KEY] = threshold_list
    eval_class_element = eval_metric_statistics_result[eval_class_type]
    eval_class_element[eval_seg_method_name] = eval_metric_statistics_element
    eval_class_element[EVAL_SEG_METHOD_NAME_KEY].append(seg_method_name)
    return eval_metric_statistics_result


def do_extract_eval_metric_statistics(eval_metric_result, eval_metric_key_list):
    threshold_list = eval_metric_result[EVAL_THRESHOLD_LIST_KEY]
    img_plane_name_list = eval_metric_result[EVAL_IMG_PLANE_NAME_LIST_KEY]
    eval_metric_result_statistics_mean = cr5_spt.init_list_storage_map(eval_metric_key_list)
    eval_metric_result_statistics_std = cr5_spt.init_list_storage_map(eval_metric_key_list)
    threshold_num = len(threshold_list)
    i = 0
    while i < threshold_num:
        threshold = threshold_list[i]
        eval_metric_result_threshold = eval_metric_result[str(threshold)]
        eval_metric_result_set = cr5_spt.init_list_storage_map(eval_metric_key_list)
        eval_metric_result_set = extract_eval_metric_set_single_threshold(eval_metric_result_threshold, img_plane_name_list, eval_metric_key_list, eval_metric_result_set)
        eval_metric_result_statistics_mean, eval_metric_result_statistics_std = fill_eval_metric_result_statistics(eval_metric_result_statistics_mean, eval_metric_result_statistics_std, eval_metric_result_set, eval_metric_key_list, threshold_list)
        i += 1
    eval_metric_result_statistics_mean = eval_metric_result_nd_array_convert(eval_metric_key_list, eval_metric_result_statistics_mean, is_metric_key_tuple=False)
    eval_metric_result_statistics_std = eval_metric_result_nd_array_convert(eval_metric_key_list, eval_metric_result_statistics_std, is_metric_key_tuple=False)
    return eval_metric_result_statistics_mean, eval_metric_result_statistics_std


def extract_eval_metric_set_single_threshold(eval_metric_result_threshold, img_plane_name_list, eval_metric_key_list, eval_metric_result_set):
    img_plane_num = len(img_plane_name_list)
    i = 0
    while i < img_plane_num:
        img_plane_name = img_plane_name_list[i]
        eval_metric_result_img_plane = eval_metric_result_threshold[img_plane_name]
        eval_metric_result_set = extract_eval_metric_set_single_plane(eval_metric_result_img_plane, eval_metric_key_list, eval_metric_result_set)
        i += 1
    return eval_metric_result_set


def extract_eval_metric_set_single_plane(eval_metric_result_img_plane, eval_metric_key_list, eval_metric_result_set):
    eval_metric_num = len(eval_metric_key_list)
    i = 0
    while i < eval_metric_num:
        eval_metric_key = eval_metric_key_list[i]
        eval_metric_result = eval_metric_result_img_plane[eval_metric_key]
        eval_metric_result_set[eval_metric_key].append(eval_metric_result)
        i += 1
    return eval_metric_result_set


def fill_eval_metric_result_statistics(eval_metric_result_statistics_mean, eval_metric_result_statistics_std, eval_metric_result_set, eval_metric_key_list, threshold_list=None):
    eval_metric_num = len(eval_metric_key_list)
    i = 0
    while i < eval_metric_num:
        eval_metric_key = eval_metric_key_list[i]
        eval_metric_result = eval_metric_result_set[eval_metric_key]
        eval_metric_result = np.array(eval_metric_result)
        eval_metric_result_mean = np.mean(eval_metric_result)
        eval_metric_result_std = np.std(eval_metric_result, axis=1)
        eval_metric_result_std = np.mean(eval_metric_result_std)
        eval_metric_result_statistics_mean[eval_metric_key].append(eval_metric_result_mean)
        eval_metric_result_statistics_std[eval_metric_key].append(eval_metric_result_std)
        i += 1
    return eval_metric_result_statistics_mean, eval_metric_result_statistics_std


def eval_img_plane_int_type_fsim(img_plane_ref, img_plane_dist):
    eval_res = imq_spt.fsim(img_plane_ref, img_plane_dist)
    return eval_res


def eval_img_plane_int_type_ssim(img_plane_ref, img_plane_dist):
    eval_res = img_qm.ssim(img_plane_ref, img_plane_dist, max_p=IMG_PLANE_MAX)
    return eval_res


def eval_img_plane_int_type_psnr(img_plane_ref, img_plane_dist):
    eval_res = img_qm.psnr(img_plane_ref, img_plane_dist, max_p=IMG_PLANE_MAX)
    return eval_res


def eval_img_plane_int_type_rmse(img_plane_ref, img_plane_dist):
    eval_res = img_qm.rmse(img_plane_ref, img_plane_dist, max_p=IMG_PLANE_MAX)
    return eval_res


EVAL_FSIM_FUN = eval_img_plane_int_type_fsim
EVAL_SSIM_FUN = eval_img_plane_int_type_ssim
EVAL_PSNR_FUN = eval_img_plane_int_type_psnr
EVAL_RMSE_FUN = eval_img_plane_int_type_rmse


def init_eval_metric_result(eval_metric_list):
    metric_list_len = len(eval_metric_list)
    eval_metric_result = dict()
    i = 0
    while i < metric_list_len:
        eval_metric = eval_metric_list[i]
        eval_metric_key = eval_metric[EVAL_METRIC_KEY_IDX]
        eval_metric_result[eval_metric_key] = []
        i += 1
    return eval_metric_result


def eval_metric_sim_cal(img_plane_int, img_plane_rec, eval_metric_list, eval_metric_result):
    metric_num = len(eval_metric_list)
    i = 0
    while i < metric_num:
        eval_metric = eval_metric_list[i]
        eval_metric_method = eval_metric[EVAL_METRIC_METHOD_IDX]
        eval_res = eval_metric_method(img_plane_int, img_plane_rec)
        eval_res = cr5_spt.nan_proc_replace(eval_res)
        eval_key = eval_metric[EVAL_METRIC_KEY_IDX]
        eval_metric_result[eval_key].append(eval_res)
        i += 1
    return eval_metric_result


def eval_metric_result_nd_array_convert(eval_metric_list, eval_metric_result, is_metric_key_tuple=True):
    metric_num = len(eval_metric_list)
    i = 0
    while i < metric_num:
        eval_key = eval_metric_list[i]
        if is_metric_key_tuple:
            eval_key = eval_key[EVAL_METRIC_KEY_IDX]
        eval_metric_result[eval_key] = np.array(eval_metric_result[eval_key])
        i += 1
    return eval_metric_result


def eval_img_plane_sim_multi_times_dynamic(img_plane, threshold_num, additional_params):
    trial_num = additional_params[EVAL_TRIAL_NUM_PARAM]
    seg_method = additional_params[EVAL_SEG_METHOD_PARAM]
    is_img_int = additional_params[IS_IMG_INT_PARAM]
    eval_metric_list = additional_params[EVAL_METRIC_LIST_PARAM]
    if is_img_int is not True:
        img_plane_int = np.round(img_plane * IMG_PLANE_MAX, 0).astype(np.uint8)
    else:
        img_plane_int = img_plane
    eval_result = init_eval_metric_result(eval_metric_list)
    eval_result[SEG_TIME_KEY] = []
    i = 0
    while i < trial_num:
        img_plane_rec, seg_time = seg_method(img_plane, threshold_num, additional_params)
        eval_result[SEG_TIME_KEY].append(seg_time)
        eval_metric_sim_cal(img_plane_int, img_plane_rec, eval_metric_list, eval_result)
        print("trial ", i, "completed")
        i += 1
    eval_result[SEG_TIME_KEY] = np.array(eval_result[SEG_TIME_KEY])
    eval_metric_result_nd_array_convert(eval_metric_list, eval_result)
    return eval_result


def get_class_type_metric_curves(eval_statistics_file_path, class_type, eval_metric):
    eval_statistics_result = cr5_spt.extract_img_objects(eval_statistics_file_path)
    eval_statistics_class_result = eval_statistics_result[class_type]
    seg_method_list = eval_statistics_class_result[EVAL_SEG_METHOD_NAME_KEY]#list(eval_statistics_class_result.keys())
    seg_method_num = len(seg_method_list)
    metric_curve_list = []
    i = 0
    while i < seg_method_num:
        seg_method_name = seg_method_list[i]
        eval_statistics_result_method = eval_statistics_class_result[seg_method_name]
        eval_statistics_value_mean = eval_statistics_result_method[EVAL_STATISTICS_MEAN_KEY]
        eval_statistics_value_mean_value = eval_statistics_value_mean[eval_metric]
        eval_statistics_value_std = eval_statistics_result_method[EVAL_STATISTICS_STD_KEY]
        eval_statistics_value_std_value = eval_statistics_value_std[eval_metric]
        eval_statistics_threshold = eval_statistics_result_method[EVAL_THRESHOLD_LIST_KEY]
        metric_curve = (seg_method_name, eval_statistics_threshold, eval_statistics_value_mean_value, eval_statistics_value_std_value)
        metric_curve_list.append(metric_curve)
        i += 1
    return metric_curve_list


def display_seg_performance_curve_mean_only(curve_list, marker_list, color_list, x_label_name, y_label_name, file_save_path=None, is_figure_show=True, title_name=None, line_width=1.5, marke_rsize=5, figure_size=(10, 6)):
    curve_num = len(curve_list)
    threshold_list_info = curve_list[0][EVAL_CURVE_THRESHOLD_IDX]
    i = 0
    #plt.ioff()
    #plt.clf()
    #plt.subplot(1, 1, 1)
    fig = plt.figure(figsize=figure_size)
    while i < curve_num:
        curve_info = curve_list[i]
        curve_value = curve_info[EVAL_CURVE_MEAN_IDX]
        plt_color = color_list[i]
        plt_marker = marker_list[i]
        plt_label = curve_info[EVAL_CURVE_NAME_IDX]
        plt.plot(threshold_list_info, curve_value, color=plt_color, marker=plt_marker, label=plt_label, linewidth=line_width, markersize=marke_rsize)
        i += 1
    plt.xlabel(x_label_name)
    plt.ylabel(y_label_name)
    if title_name is not None:
        plt.title(title_name)
    plt.legend()
    if file_save_path is not None:
        plt.savefig(file_save_path, bbox_inches="tight")
    if is_figure_show:
       plt.show()
    plt.close(fig)
    return


def output_eval_metric_table_query_results_files_to_xlsx(eval_metric_query_result_file_path_list, out_put_file_path):
    file_num = len(eval_metric_query_result_file_path_list)
    eval_metric_query_result_list = []
    i = 0
    while i < file_num:
        eval_metric_query_result_file_path = eval_metric_query_result_file_path_list[i]
        eval_metric_query_result = cr5_spt.extract_img_objects(eval_metric_query_result_file_path)
        eval_metric_query_result_list.append(eval_metric_query_result)
        i += 1
    output_eval_metric_table_query_results_to_xlsx(eval_metric_query_result_list, out_put_file_path)
    return


def output_eval_metric_table_query_results_to_xlsx(eval_metric_query_result_list, out_put_file_path):
    workbook = Workbook()
    eval_metric_query_result_num = len(eval_metric_query_result_list)
    i = 0
    while i < eval_metric_query_result_num:
        eval_metric_query_result = eval_metric_query_result_list[i]
        output_eval_metric_table_query_result_to_xlsx_sheet(workbook, eval_metric_query_result)
        i += 1
    clear_class_metric_work_book(workbook)
    workbook.save(out_put_file_path)
    return


def output_eval_metric_table_query_result_to_xlsx_sheet(workbook, eval_metric_query_result):
    class_type = eval_metric_query_result[EVAL_METRIC_TABLE_QUERY_RESULT_CLASS_TYPE_IDX]
    metric_type = eval_metric_query_result[EVAL_METRIC_TABLE_QUERY_METRIC_TYPE_IDX]
    metric_seg_method = eval_metric_query_result[EVAL_METRIC_TABLE_QUERY_SEG_METHOD_IDX]
    threshold = eval_metric_query_result[EVAL_METRIC_TABLE_QUERY_THRESHOLD_NUM_IDX]
    image_plane_name = eval_metric_query_result[EVAL_METRIC_TABLE_QUERY_IMAGE_PLANE_IDX]
    eval_query_metric = eval_metric_query_result[EVAL_METRIC_TABLE_QUERY_METRIC_VALUE_IDX]
    worksheet = create_class_metric_work_sheet(workbook, class_type, metric_type)
    worksheet = insert_xlsx_row_wise(worksheet, EVAL_METRIC_TABLE_QUERY_METRIC_SEG_METHOD_ROW, EVAL_METRIC_TABLE_QUERY_METRIC_SEG_METHOD_COLUMN_START, metric_seg_method, is_cell_title=True)
    threshold_num = len(threshold)
    image_plane_num = len(image_plane_name)
    i = 0
    k = 0
    while i < threshold_num:
        threshold_value = threshold[i]
        threshold_start_row = k + EVAL_METRIC_XLS_DATA_ROW_START
        threshold_end_row = threshold_start_row + image_plane_num - 1
        insert_xlsx_row_wise(worksheet, row=threshold_start_row, column_start=EVAL_METRIC_TABLE_QUERY_METRIC_THRESHOLD_COLUMN, value_list=[threshold_value], is_cell_title=True)
        j = 0
        while j < image_plane_num:
            image_plane_name_value = proc_image_plane_name(image_plane_name[j], is_contain_ext=False)
            insert_xlsx_row_wise(worksheet, row=k + EVAL_METRIC_XLS_DATA_ROW_START, column_start=EVAL_METRIC_TABLE_QUERY_METRIC_IMAGE_PLANE_NAME_COLUMN, value_list=[image_plane_name_value], is_cell_title=True)
            query_metric_value = eval_query_metric[k]
            query_metric_mean_value = query_metric_value[EVAL_METRIC_TABLE_QUERY_METRIC_MEAN_VALUE_IDX]
            insert_xlsx_row_wise(worksheet, row=k + EVAL_METRIC_XLS_DATA_ROW_START, column_start=EVAL_METRIC_TABLE_QUERY_METRIC_SEG_METHOD_COLUMN_START, value_list=query_metric_mean_value)
            j += 1
            k += 1
        worksheet.merge_cells(start_row=threshold_start_row, start_column=EVAL_METRIC_TABLE_QUERY_METRIC_THRESHOLD_COLUMN, end_row=threshold_end_row, end_column=EVAL_METRIC_TABLE_QUERY_METRIC_THRESHOLD_COLUMN)
        i += 1
    return worksheet


def proc_image_plane_name(image_plane_name, is_contain_ext=True):
    if not is_contain_ext:
        image_plane_name = image_plane_name.split(FILE_NAME_EXT_SPLITER)[0]
    return image_plane_name


def create_class_metric_work_sheet(workbook, class_type, metric_type):
    sheet_name = class_type + EVAL_METRIC_TABLE_QUERY_SHEET_NAME_CONN + metric_type
    worksheet = workbook.create_sheet(sheet_name)
    set_cell_title_style(worksheet.cell(row=EVAL_METRIC_XLS_TITLE_ROW, column=EVAL_METRIC_XLS_TITLE_COL_1, value=EVAL_METRIC_XLS_THRESHOLD_NUM_TITLE))
    set_cell_title_style(worksheet.cell(row=EVAL_METRIC_XLS_TITLE_ROW, column=EVAL_METRIC_XLS_TITLE_COL_2, value=EVAL_METRIC_XLS_IMAGE_NAME_TITLE))
    return worksheet


def set_cell_title_style(cell_element, is_border_only=False):
    cell_element.border = CELL_BOARDER
    if not is_border_only:
        cell_element.font = Font(name='Calibri', size=11, bold=True, italic=False)
        cell_element.fill = PatternFill(fill_type='solid', start_color='E0E0E0', end_color='E0E0E0')
        cell_element.alignment = Alignment(horizontal='center', vertical='center')
    return cell_element


def clear_class_metric_work_book(workbook):
    try:
        worksheet_initial = workbook[XLS_SHEET_NAME_INITIAL]
        workbook.remove(worksheet_initial)
    except:
        print(XLS_SHEET_INITIAL_NO_FOUND_MSG)
    return workbook


def insert_xlsx_row_wise(worksheet, row, column_start, value_list, is_cell_title=False):
    i = 0
    value_num = len(value_list)
    while i < value_num:
        cell_value = value_list[i]
        cell_element = worksheet.cell(row=row, column=column_start+i, value=cell_value)
        if is_cell_title:
            set_cell_title_style(cell_element)
        else:
            set_cell_title_style(cell_element, is_border_only=True)
        i += 1
    return worksheet
### 增加置信区间画图代码
def display_seg_performance_curve_with_std(curve_list, marker_list, color_list, x_label_name, y_label_name,
                                           file_save_path=None, is_figure_show=True, title_name=None,
                                           line_width=1.5, marke_rsize=7, figure_size=(10, 6), alpha=0.15):
    """
    绘制带有标准差阴影区域的性能曲线
    alpha: 阴影部分的透明度
    """
    curve_num = len(curve_list)
    fig = plt.figure(figsize=figure_size)

    for i in range(curve_num):
        curve_info = curve_list[i]

        # 提取数据
        label_name = curve_info[EVAL_CURVE_NAME_IDX]  # 下标 0
        thresholds = curve_info[EVAL_CURVE_THRESHOLD_IDX]  # 下标 1
        means = np.array(curve_info[EVAL_CURVE_MEAN_IDX])  # 下标 2
        stds = np.array(curve_info[EVAL_CURVE_STD_IDX])  # 下标 3

        plt_color = color_list[i]
        plt_marker = marker_list[i]

        # 1. 绘制主线（均值）
        plt.plot(thresholds, means, color=plt_color, marker=plt_marker,
                 label=label_name, linewidth=line_width, markersize=marke_rsize)

        # 2. 绘制置信区间（均值 ± 标准差）
        # 使用 fill_between 填充阴影
        plt.fill_between(thresholds, means - stds, means + stds,
                         color=plt_color, alpha=alpha, label='_nolegend_')

    plt.xlabel(x_label_name)
    plt.ylabel(y_label_name)
    if title_name is not None:
        plt.title(title_name)
    plt.legend()
    plt.grid(True, linestyle='--', alpha=0.6)  # 增加网格线使图表更专业

    if file_save_path is not None:
        plt.savefig(file_save_path, bbox_inches="tight")
    if is_figure_show:
        plt.show()
    plt.close(fig)